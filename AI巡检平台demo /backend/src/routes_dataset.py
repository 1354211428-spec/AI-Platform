"""
数据库模块路由 - 数据集管理、Excel上传、SQL取数
"""
import os
import json
import openpyxl
from flask import Blueprint, request, jsonify
from .database import get_db
from .utils import success, error, rows_to_list, row_to_dict, parse_json_field, now_str

dataset_bp = Blueprint('dataset', __name__)
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), '..', 'uploads')


# ==================== 数据集列表 ====================
@dataset_bp.route('/datasets', methods=['GET'])
def list_datasets():
    source_type = request.args.get('source_type', '')
    keyword = request.args.get('keyword', '')
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 20))
    offset = (page - 1) * page_size

    conn = get_db()
    cursor = conn.cursor()

    where_clauses = []
    params = []
    if source_type:
        where_clauses.append("source_type = ?")
        params.append(source_type)
    if keyword:
        where_clauses.append("name LIKE ?")
        params.append(f"%{keyword}%")

    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    cursor.execute(f"SELECT COUNT(*) FROM dataset {where_sql}", params)
    total = cursor.fetchone()[0]

    cursor.execute(f"SELECT * FROM dataset {where_sql} ORDER BY created_at DESC LIMIT ? OFFSET ?",
                   params + [page_size, offset])
    rows = rows_to_list(cursor.fetchall())
    conn.close()

    for r in rows:
        r['field_schema'] = parse_json_field(r.get('field_schema'))

    return jsonify(success({
        "list": rows,
        "total": total,
        "page": page,
        "page_size": page_size
    }))


# ==================== 数据集详情 ====================
@dataset_bp.route('/datasets/<int:dataset_id>', methods=['GET'])
def get_dataset(dataset_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM dataset WHERE id = ?", (dataset_id,))
    row = row_to_dict(cursor.fetchone())
    conn.close()
    if not row:
        return jsonify(error("数据集不存在", 404)), 404
    row['field_schema'] = parse_json_field(row.get('field_schema'))
    return jsonify(success(row))


# ==================== Excel 上传 ====================
@dataset_bp.route('/datasets/upload', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return jsonify(error("请上传文件")), 400

    file = request.files['file']
    name = request.form.get('name', file.filename.rsplit('.', 1)[0])

    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        return jsonify(error("仅支持 .xlsx .xls .csv 格式")), 400

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(file_path)

    # 解析Excel获取字段信息
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        headers = [str(cell.value) for cell in next(ws.iter_rows(max_row=1)) if cell.value]
        record_count = ws.max_row - 1  # 减去表头行
        wb.close()

        # 自动推断字段类型
        field_schema = []
        for h in headers:
            ft = "text"
            h_lower = h.lower()
            if any(k in h_lower for k in ['id', 'count', 'num', 'price', 'score']):
                ft = "number"
            elif any(k in h_lower for k in ['url', 'link', 'img', 'image']):
                ft = "url"
            elif any(k in h_lower for k in ['date', 'time', 'at', '_dt']):
                ft = "date"
            elif any(k in h_lower for k in ['tag', 'category', 'type', 'label']):
                ft = "tag"
            field_schema.append({"name": h, "type": ft})

    except Exception as e:
        return jsonify(error(f"文件解析失败: {str(e)}")), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO dataset (name, source_type, field_schema, record_count, status, file_path, created_by)
        VALUES (?, 'excel', ?, ?, 'ready', ?, 'admin')
    """, (name, json.dumps(field_schema), record_count, file_path))
    dataset_id = cursor.lastrowid
    conn.commit()

    cursor.execute("SELECT * FROM dataset WHERE id = ?", (dataset_id,))
    row = row_to_dict(cursor.fetchone())
    conn.close()

    row['field_schema'] = parse_json_field(row.get('field_schema'))
    return jsonify(success(row))


# ==================== SQL 取数 ====================
@dataset_bp.route('/datasets/sql', methods=['POST'])
def create_sql_dataset():
    data = request.get_json()
    name = data.get('name', '未命名数据集')
    sql_content = data.get('sql_content', '')

    if not sql_content.strip():
        return jsonify(error("SQL内容不能为空")), 400

    # 模拟字段解析（Demo: 从 SELECT 子句解析字段名）
    field_schema = _parse_sql_fields(sql_content)

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO dataset (name, source_type, field_schema, record_count, status, sql_content, created_by)
        VALUES (?, 'sql', ?, ?, 'ready', ?, 'admin')
    """, (name, json.dumps(field_schema), 0, sql_content))
    dataset_id = cursor.lastrowid
    conn.commit()

    cursor.execute("SELECT * FROM dataset WHERE id = ?", (dataset_id,))
    row = row_to_dict(cursor.fetchone())
    conn.close()

    row['field_schema'] = parse_json_field(row.get('field_schema'))
    return jsonify(success(row))


# ==================== SQL 验证 ====================
@dataset_bp.route('/sql/validate', methods=['POST'])
def validate_sql():
    data = request.get_json()
    sql = data.get('sql', '')
    # Demo: 简单规则校验
    errors = []
    if not sql.strip().upper().startswith('SELECT'):
        errors.append("SQL必须以 SELECT 开头")
    if 'DROP' in sql.upper() or 'DELETE' in sql.upper() or 'TRUNCATE' in sql.upper():
        errors.append("不允许使用 DROP/DELETE/TRUNCATE 语句")

    if errors:
        return jsonify(success({"valid": False, "errors": errors}))
    return jsonify(success({"valid": True, "errors": []}))


# ==================== SQL 资源预估 ====================
@dataset_bp.route('/sql/estimate', methods=['POST'])
def estimate_sql():
    import random
    data = request.get_json()
    sql = data.get('sql', '')
    # Demo: 模拟预估
    estimated_rows = random.randint(1000, 50000)
    estimated_seconds = random.randint(30, 300)
    return jsonify(success({
        "estimated_rows": estimated_rows,
        "estimated_seconds": estimated_seconds,
        "estimated_size_mb": round(estimated_rows * 0.002, 1),
        "warning": "数据量较大，建议增加分区过滤" if estimated_rows > 30000 else None
    }))


# ==================== 数据预览 ====================
@dataset_bp.route('/datasets/<int:dataset_id>/preview', methods=['GET'])
def preview_dataset(dataset_id):
    import random
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM dataset WHERE id = ?", (dataset_id,))
    ds = row_to_dict(cursor.fetchone())
    conn.close()

    if not ds:
        return jsonify(error("数据集不存在", 404)), 404

    field_schema = parse_json_field(ds.get('field_schema')) or []

    # 生成模拟预览数据
    preview_rows = []
    for i in range(min(10, ds.get('record_count', 10))):
        row = {}
        for f in field_schema:
            fname = f['name']
            ftype = f['type']
            if ftype == 'number':
                row[fname] = random.randint(10000, 99999)
            elif ftype == 'url':
                row[fname] = f"https://cdn.kuaishou.com/item/{random.randint(1000, 9999)}.jpg"
            elif ftype == 'date':
                row[fname] = f"2024-0{random.randint(1, 9)}-{random.randint(10, 28)}"
            elif ftype == 'tag':
                row[fname] = random.choice(["服装", "鞋包", "3C数码", "家居", "美妆"])
            else:
                row[fname] = f"示例数据_{i + 1}"
        preview_rows.append(row)

    return jsonify(success({
        "fields": field_schema,
        "rows": preview_rows,
        "total": ds.get('record_count', 0)
    }))


# ==================== 删除数据集 ====================
@dataset_bp.route('/datasets/<int:dataset_id>', methods=['DELETE'])
def delete_dataset(dataset_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM dataset WHERE id = ?", (dataset_id,))
    conn.commit()
    conn.close()
    return jsonify(success({"id": dataset_id}))


def _parse_sql_fields(sql):
    """简单解析SQL SELECT字段"""
    import re
    sql_upper = sql.upper()
    try:
        select_start = sql_upper.index('SELECT') + 6
        from_start = sql_upper.index('FROM')
        select_clause = sql[select_start:from_start].strip()
        if select_clause == '*':
            return [{"name": "field_1", "type": "text"}]
        fields = []
        for part in select_clause.split(','):
            part = part.strip()
            # 取 AS 别名或最后一个词
            if ' AS ' in part.upper():
                name = re.split(r'\s+as\s+', part, flags=re.IGNORECASE)[-1].strip()
            else:
                name = part.split('.')[-1].strip()
            # 清理特殊字符
            name = re.sub(r'[^a-zA-Z0-9_\u4e00-\u9fa5]', '', name)
            if name:
                ft = "text"
                n = name.lower()
                if any(k in n for k in ['id', 'count', 'num', 'price']):
                    ft = "number"
                elif any(k in n for k in ['date', 'time', 'dt']):
                    ft = "date"
                fields.append({"name": name, "type": ft})
        return fields if fields else [{"name": "field_1", "type": "text"}]
    except:
        return [{"name": "field_1", "type": "text"}]
